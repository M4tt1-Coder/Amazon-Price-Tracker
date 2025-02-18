import openpyxl
import pandas as pd
from .get_productdata import get_data_np
import os

def receive_data_np(url,file):
    # receive 5 return values from 'get_data_np()'
    price, date, name, description, id = get_data_np(url)
    # create dictionary for return values
    product_entry = {"price": price,
                     "date": date,
                     "name": name,
                     "description": description,
                     "id": id}

    file_exists = os.path.exists(file)

    # load existing data if the file exists
    if file_exists:
        with pd.ExcelFile(file, engine='openpyxl') as excel_data:
            sheet_exists = id in excel_data.sheet_names
            existing_data = pd.read_excel(file, sheet_name=id) if sheet_exists else pd.DataFrame()
    else:
        sheet_exists = False
        existing_data = pd.DataFrame()

    # create DataFrame for new product
    new_data = pd.DataFrame([product_entry])

    # combine old and new data if the sheet exists
    updated_data = pd.concat([existing_data, new_data], ignore_index=True) if sheet_exists else new_data

    # save to excel, replacing the sheet if it exists
    with pd.ExcelWriter(file, engine='openpyxl', mode='a' if file_exists else 'w', if_sheet_exists='replace' if sheet_exists else None) as writer:
        updated_data.to_excel(writer, sheet_name=id, index=False)

    print(f"Added: {product_entry}")

def delete_excel_sheet(file, sheet_name):
    """Deletes a specific sheet from an Excel file."""
    if not os.path.exists(file):
        print(f"File '{file}' does not exist.")
        return

    wb = openpyxl.load_workbook(file)

    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found in '{file}'.")
        return

    # remove the excel-sheet
    wb.remove(wb[sheet_name])

    # if no sheets remain, delete the entire file
    if not wb.sheetnames:
        os.remove(file)
        print(f"Deleted last sheet '{sheet_name}' and removed file '{file}'.")
    else:
        wb.save(file)
        print(f"Deleted sheet '{sheet_name}' from '{file}'.")

def delete_all_products(file):
    if os.path.exists(file):
        os.remove(file)
        print(f"Removed file '{file}'.")
    else:
        print(f"File '{file}' does not exist.")
